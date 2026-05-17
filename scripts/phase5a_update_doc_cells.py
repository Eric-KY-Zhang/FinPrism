"""Phase 5a — rewrite user-facing copy that still mentions the manual
xq_a_token in 样本池!E5.

Usage (run AFTER you've backed up the .xlsm and closed it in Excel):

    py scripts/phase5a_update_doc_cells.py --dry-run     # default, prints planned edits
    py scripts/phase5a_update_doc_cells.py --apply       # writes in place

The script is defensive on purpose. The cells named in PHASE5A spec are
the intended targets, but the current workbook layout has drifted (no
'使用说明' sheet; the 使用提示 panel anchor moved off S2). We:

  1. Always rename 样本池!A5 anchor label.
  2. Walk every visible sheet to find any cell whose text still talks
     about pasting the cookie into E5 / B5, and replace inline.
  3. Update 使用说明 §4 / §5 / FAQ rows ONLY if the sheet exists.
  4. Print every planned edit; in --dry-run mode no file is touched.

Nothing in this script should be considered binding if the printed
preview looks wrong — abort, hand-edit, and rerun.
"""
from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "上市公司财务数据查询.xlsm"


# (sheet, cell, original_substring, new_text_full)
# original_substring is matched against current value; if missing we skip.
INTRO_REPLACEMENTS = [
    (
        "使用说明",
        "D29",
        "雪球",
        "Phase 5a 起雪球抓数已改为匿名 session 自动 warmup,无需在 E5 维护 cookie。E5 留空即可。",
    ),
    (
        "使用说明",
        "C48",
        "雪球港股财务数据",
        "雪球港股财务数据。Phase 5a 起匿名 warmup 自动获取 session cookie,E5 留空。",
    ),
    (
        "使用说明",
        "D65",
        "雪球",
        "若批量失败,先检查诊断 sheet 的 HTTP 状态码。匿名 session 限流可能比登录态严,出现 429/403 需调大节流。",
    ),
]


def _norm(val) -> str:
    return "" if val is None else str(val)


def patch_pool_label(ws, planned: list[tuple[str, str, str, str]]) -> None:
    """样本池!A5 label rename."""
    cell = ws["A5"]
    before = _norm(cell.value)
    after = "雪球 Cookie (已弃用,留空)"
    if before == after:
        return
    if "Cookie" not in before and "cookie" not in before:
        print(f"  ! 样本池!A5 doesn't look like the cookie label: {before!r} — skipping")
        return
    planned.append(("样本池", "A5", before, after))


def sweep_cookie_hints(wb, planned: list[tuple[str, str, str, str]]) -> None:
    """Find any cell text that still tells the user to paste a cookie into E5/B5."""
    needles = (
        "E5 填雪球 cookie",
        "E5 填雪球Cookie",
        "B5 填雪球 cookie",
        "B5 填雪球Cookie",
        "xq_a_token",
        "雪球 cookie",
    )
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:  # type: Cell
                v = cell.value
                if not isinstance(v, str) or len(v) > 4000:
                    continue
                hit = next((n for n in needles if n in v), None)
                if hit is None:
                    continue
                rewritten = v
                rewritten = rewritten.replace(
                    "2. E5 填雪球 cookie;E6 切原币 / 统一RMB",
                    "2. E6 切原币 / 统一RMB",
                )
                rewritten = rewritten.replace(
                    "2. B5 填雪球 cookie;E6 切原币 / 统一RMB",
                    "2. E6 切原币 / 统一RMB",
                )
                rewritten = rewritten.replace(
                    "E5 填雪球 cookie",
                    "(E5 已弃用,Phase 5a 起匿名 session)",
                )
                rewritten = rewritten.replace(
                    "B5 填雪球 cookie",
                    "(E5 已弃用,Phase 5a 起匿名 session)",
                )
                # Heuristic: if we changed step-2 wording, also re-number 3/4/5 down by one.
                if "2. E6 切原币" in rewritten and "3. 第 14 行" in rewritten:
                    rewritten = rewritten.replace("\n3. ", "\n3. ", 1)  # placeholder, edit below
                if rewritten != v:
                    planned.append((ws.title, cell.coordinate, v, rewritten))


def patch_intro_sheet(wb, planned: list[tuple[str, str, str, str]]) -> None:
    if "使用说明" not in wb.sheetnames:
        print("  ! '使用说明' sheet not found in this workbook — skipping intro D29/C48/D65 edits.")
        return
    ws = wb["使用说明"]
    for sheet_name, addr, marker, new_text in INTRO_REPLACEMENTS:
        cell = ws[addr]
        v = _norm(cell.value)
        if not v:
            print(f"  ! 使用说明!{addr} empty — skipping (layout may have shifted)")
            continue
        if marker not in v:
            print(f"  ! 使用说明!{addr} doesn't contain marker {marker!r}; current = {v[:80]!r} — skipping")
            continue
        if v == new_text:
            continue
        planned.append(("使用说明", addr, v, new_text))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write changes to the .xlsm in place (default: dry-run)")
    parser.add_argument("--book", default=str(BOOK), help="Path to .xlsm (default: 上市公司财务数据查询.xlsm at repo root)")
    parser.add_argument("--backup", action="store_true", help="Copy <book>.bak before --apply")
    args = parser.parse_args()

    book = Path(args.book)
    if not book.exists():
        print(f"FATAL: book not found: {book}")
        return 2

    wb = load_workbook(str(book), keep_vba=True, data_only=False)
    if "样本池" not in wb.sheetnames:
        print("FATAL: 样本池 sheet missing — aborting")
        return 2

    planned: list[tuple[str, str, str, str]] = []
    patch_pool_label(wb["样本池"], planned)
    sweep_cookie_hints(wb, planned)
    patch_intro_sheet(wb, planned)

    print(f"\n=== Planned edits ({len(planned)}) ===")
    for sheet, addr, before, after in planned:
        print(f"  {sheet}!{addr}")
        print(f"    BEFORE: {before[:160]!r}")
        print(f"    AFTER : {after[:160]!r}")

    if not planned:
        print("Nothing to do.")
        return 0

    if not args.apply:
        print("\n--dry-run (no file written). Re-run with --apply to commit.")
        return 0

    if args.backup:
        backup = book.with_suffix(book.suffix + ".bak")
        shutil.copy2(book, backup)
        print(f"\nBacked up to {backup}")

    for sheet, addr, _before, after in planned:
        wb[sheet][addr] = after
    wb.save(str(book))
    print(f"\nWrote {len(planned)} edits to {book}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
