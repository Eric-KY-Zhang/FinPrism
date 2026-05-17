"""Read-only probe: inspect 使用说明 / 样本池 cells we plan to rewrite in
phase5a_update_doc_cells.py, so we can verify the cell addresses still hold
the assumed content before mutating them.

Run with:
    py tools/phase5a_inspect_doc_cells.py
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "上市公司财务数据查询.xlsm"


def show(ws, addr: str) -> None:
    val = ws[addr].value
    print(f"  {ws.title}!{addr} = {val!r}")


def main() -> int:
    wb = load_workbook(str(BOOK), keep_vba=True, data_only=False)
    print("=== Sheets (with visibility) ===")
    for name in wb.sheetnames:
        ws_i = wb[name]
        print(f"  - {name}  (state={ws_i.sheet_state})")
    if "使用说明" in wb.sheetnames:
        ws = wb["使用说明"]
        print("\n=== 使用说明 — target cells ===")
        for addr in ("D29", "C48", "D65"):
            show(ws, addr)
    else:
        print("WARN: 使用说明 sheet missing")
    if "样本池" in wb.sheetnames:
        ws2 = wb["样本池"]
        print("\n=== 样本池 — target cells ===")
        for addr in ("A5", "B5", "E5", "S2", "W2"):
            show(ws2, addr)
    else:
        print("WARN: 样本池 sheet missing")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
