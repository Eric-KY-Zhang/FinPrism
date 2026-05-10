"""
上市公司财务数据查询 — 工作簿模板生成器

生成空的 xlsm 模板,含使用说明、样本池、A股/美股/港股/韩股 16 张正式表的结构 + 列宽 + 表头容器 + 冻结窗格。
不含任何 VBA 代码 — 后续由 install_modules.py 注入 modules/*.bas。

用法:
    cd "E:\\Claude+CODEX Project\\FS Capture\\VBA Captor"
    py tools/build_template.py

输出:
    上市公司财务数据查询.xlsx  (在工作目录根)

注意:
    输出 .xlsx 而非 .xlsm — openpyxl 给 fresh 工作簿写 .xlsm 时不带 vbaProject,
    Excel 会拒绝打开。后续由 tools/install_modules.py 用 Excel COM 转成 .xlsm
    并注入 VBA 模块。
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_PATH = Path(__file__).resolve().parent.parent / "上市公司财务数据查询.xlsx"

DARK_BLUE = "FF4472C4"
SECONDARY_BLUE = "FFD9E1F2"
SECONDARY_FG = "FF1F4E79"
LIGHT_GRAY = "FFD9D9D9"
WHITE = "FFFFFFFF"

THIN = Side(border_style="thin", color="FF808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
SUB_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FF000000")
DATA_FONT = Font(name="微软雅黑", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_intro(ws):
    ws.column_dimensions["A"].width = 100
    ws["A1"] = "上市公司财务数据查询"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True)

    lines = [
        "",
        "【用途】把上市公司财务数据抓成同业对标宽表,方便横向比较。",
        "【当前支持】A股、美股、港股、韩股。",
        "【后续规划】更多市场。",
        "【作者】Eric Zhang",
        "【联系邮箱】214978902@qq.com",
        "",
        "【使用步骤】",
        "1. 在『样本池』Sheet 第 14 行起按市场录入公司:",
        "     A:B=A股, D:E=美股, G:H=港股, J:K=韩股; 每个市场只填代码和简称。",
        "2. A2 填年份 (如 2025), 留空=取最新可用期间。",
        "3. A4 选择季度: 全部 / Q1 / Q2 / Q3 / Q4。",
        "4. B5 可填写雪球 xq_a_token cookie; POM、HTT 等 EDGAR 不完整的中概/20-F 公司会自动走雪球 fallback, 港股也使用该 cookie。",
        "5. 点样本池顶部按钮抓数:",
        "     【一键 A股 / 一键 美股 / 一键 港股 / 一键 韩股】— 只更新对应市场 4 张表",
        "     【一键全抓 4 市场】— 顺序更新 A股、美股、港股、韩股 16 张表",
        "     16 个单表按钮保留在样本池下方辅助区,用于单独排查。",
        "",
        "【宽表格式】",
        "  R1: 公司名(代码), 跨该公司所有报告期合并",
        "  R2: 报告期, 降序排列",
        "  A/B列: 大类或指标类型、指标名称; 指标表额外有 C列英文指标名",
        "  A股: 报告期跨公司取并集对齐; 美股/港股/韩股: 每家公司只展开自己有数据的报告期",
        "  港股: 单位为百万(各家公司报告币种, 见 港股_抓取诊断 Unit 列); 默认原币输出",
        "  韩股: 单位为十亿韩元(KRW billions), 数据源表格为百万韩元, 写表时除以 1,000",
        "",
        "【数据源】",
        "  A股: 新浪财经",
        "  美股: SEC EDGAR companyfacts; 中概/20-F fallback 到雪球",
        "  港股: 雪球 HK API",
        "  韩股: stockanalysis.com KRX 财报 HTML 表格",
        "",
        "【限制】",
        "  - 雪球 cookie 过期时需重新复制 xq_a_token 到 B5",
        "  - 诊断 sheet 中同一 (公司, 指标) 先出现 MISSING_NON_USD、随后出现 OK_XUEQIU 属预期行为:表示 ifrs-full 有字段但单位不是 USD,系统改走雪球兜底",
        "  - 样本池已按市场分栏,韩股代码请填在韩股区,不再需要市场列",
        "",
        "【汇率与币种】",
        "  新增『汇率』sheet 缓存 USDCNY / HKDCNY / KRWCNY 期末与期间平均汇率。",
        "  数据源: 雪球 K 线 USDCNY.FX / HKDCNY.FX / KRWCNY.FX, 期间平均 = 区间内日 close 算术平均。",
        "  1. 在样本池 A 列填代码、B 列填简称 (各市场分栏)。",
        "  2. B5 填雪球 xq_a_token cookie (港股抓数 / 美股雪球 fallback 使用; 汇率缓存可自动 warmup)。",
        "  3. B6 选 '原币' (默认) 或 '统一RMB' (4 市场全部按当期汇率换算成 RMB 显示)。",
        "  4. B8 默认 '关';仅当 EDGAR + 雪球均失败时,可手动设 '开' 启用 BABA/JD/PDD stockanalysis 备用路径。",
        "  5. 点 '一键全抓 4 市场', 等候 ~3 分钟。",
        "  6. 切换 B6 后已写表数值会立即按公式切换显示,无需重新抓数。",
        "  汇率值在『汇率』sheet 缓存;HTTP 响应缓存写入 .cache/ 24 小时;均可本地清理或手动 override。",
        "",
        "【来源说明】基于林铖 V2.2 重写并扩展。",
    ]
    for i, text in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = LEFT
        cell.font = Font(name="微软雅黑", size=11)


def build_sample_pool(ws):
    """
    样本池布局 (Phase 4e):
      Row 1-5: 全局配置区
      Row 7: 4 市场标题
      Row 8: 4 市场一键按钮位
      Row 9: 4 市场 tabs 显隐按钮位
      Row 10: 各市场列头
      Row 11+: 公司数据
    """
    sub_header_fill = PatternFill("solid", fgColor="FFB4C7E7")    # 浅蓝
    yellow_fill = PatternFill("solid", fgColor="FFFFE699")
    cookie_fill = PatternFill("solid", fgColor="FFFFF2CC")

    # ---- 列宽 ----
    col_widths = {
        "A": 11, "B": 16, "C": 2, "D": 2,
        "E": 8, "F": 18, "G": 2, "H": 2,
        "I": 7, "J": 14, "K": 2, "L": 2,
        "M": 8, "N": 16, "O": 2, "P": 2,
        "Q": 22, "R": 2, "S": 24,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ---- Row 1-6: 全局配置区 ----
    for addr, txt in (
        ("A1", "年份 (留空=取最新)"),
        ("A3", "季度 (Q1/Q2/Q3/Q4 或 全部)"),
        ("A5", "雪球 Cookie"),
        ("A6", "显示币种"),
    ):
        ws[addr] = txt
        ws[addr].font = Font(name="微软雅黑", size=10, bold=True)
        ws[addr].fill = sub_header_fill
        ws[addr].alignment = CENTER
        ws[addr].border = BORDER

    ws["A2"] = 2025
    ws["A2"].font = Font(name="微软雅黑", size=11, bold=True)
    ws["A2"].fill = yellow_fill
    ws["A2"].alignment = CENTER
    ws["A2"].border = BORDER

    ws["A4"] = "全部"
    ws["A4"].font = Font(name="微软雅黑", size=11, bold=True)
    ws["A4"].fill = yellow_fill
    ws["A4"].alignment = CENTER
    ws["A4"].border = BORDER
    dv = DataValidation(type="list", formula1='"全部,Q1,Q2,Q3,Q4"', allow_blank=False)
    dv.add("A4")
    ws.add_data_validation(dv)

    ws.merge_cells("B5:F5")
    ws["B5"] = ""
    ws["B5"].font = Font(name="Consolas", size=9)
    ws["B5"].fill = cookie_fill
    ws["B5"].alignment = LEFT

    # Phase 4f Step 2: B6 显示币种 toggle (默认 "原币", 下拉 原币/统一RMB)
    ws["B6"] = "原币"
    ws["B6"].font = Font(name="微软雅黑", size=11, bold=True)
    ws["B6"].fill = yellow_fill
    ws["B6"].alignment = CENTER
    ws["B6"].border = BORDER
    dv_currency = DataValidation(type="list", formula1='"原币,统一RMB"', allow_blank=False)
    dv_currency.add("B6")
    ws.add_data_validation(dv_currency)

    for row in range(1, 7):
        ws.row_dimensions[row].height = 22

    # ---- Row 7: 市场标题 ----
    markets = [
        ("A7:B7", "A 股(新浪)", "FF4472C4"),
        ("E7:F7", "美股(EDGAR+雪球)", "FFC00000"),
        ("I7:J7", "港股(雪球 HK)", "FF548235"),
        ("M7:N7", "韩股(stockanalysis)", "FF7030A0"),
    ]
    for rng, title, fill in markets:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = title
        c.font = HEADER_FONT
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = CENTER
        c.border = BORDER

    ws["B8"] = "关"
    ws["B8"].font = Font(name="微软雅黑", size=9, bold=True)
    ws["B8"].fill = yellow_fill
    ws["B8"].alignment = CENTER
    ws["B8"].border = BORDER
    dv_sa = DataValidation(type="list", formula1='"关,开"', allow_blank=False)
    dv_sa.add("B8")
    ws.add_data_validation(dv_sa)

    # ---- Row 8-9 / Q1-Q10: 按钮位提示, install_modules.py 会覆盖为 Shape 按钮 ----
    button_placeholders = [
        ("A8:A8", "一键 A 股", "FF4472C4", WHITE, 10),
        ("E8:F8", "一键 美股", "FFC00000", WHITE, 11),
        ("I8:J8", "一键 港股", "FF548235", WHITE, 11),
        ("M8:N8", "一键 韩股", "FF7030A0", WHITE, 11),
        ("Q1:Q3", "一键全抓 4 市场", "FF4472C4", WHITE, 11),
        ("Q5:Q7", "合并跨市场指标表", "FF4472C4", WHITE, 11),
        ("Q14:Q14", "清空缓存", SECONDARY_BLUE, SECONDARY_FG, 9),
        ("S1:S3", "合并 4 张跨市场表", "FF4472C4", WHITE, 11),
        ("S5:S7", "合并跨市场资产负债表", "FF4472C4", WHITE, 10),
        ("S8:S10", "合并跨市场利润表", "FF4472C4", WHITE, 10),
        ("S11:S13", "合并跨市场现金流量表", "FF4472C4", WHITE, 10),
        ("A9:B9", "切换 A 股 tabs 显隐", SECONDARY_BLUE, SECONDARY_FG, 9),
        ("E9:F9", "切换 美股 tabs 显隐", SECONDARY_BLUE, SECONDARY_FG, 9),
        ("I9:J9", "切换 港股 tabs 显隐", SECONDARY_BLUE, SECONDARY_FG, 9),
        ("M9:N9", "切换 韩股 tabs 显隐", SECONDARY_BLUE, SECONDARY_FG, 9),
        ("Q8:Q10", "切换所有分市场 tabs 显隐", SECONDARY_BLUE, SECONDARY_FG, 10),
    ]
    for rng, title, fill, font_color, font_size in button_placeholders:
        if rng.split(":")[0] != rng.split(":")[-1]:
            ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = title
        c.font = Font(name="微软雅黑", size=font_size, bold=True, color=font_color)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = CENTER
        c.border = BORDER

    # ---- Row 10: 数据表头 ----
    for code_col, name_col in (("A", "B"), ("E", "F"), ("I", "J"), ("M", "N")):
        for col, name in ((code_col, "代码"), (name_col, "简称")):
            cell = ws[f"{col}10"]
            cell.value = name
            cell.font = HEADER_FONT
            cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
            cell.alignment = CENTER
            cell.border = BORDER

    ws.row_dimensions[7].height = 24
    ws.row_dimensions[8].height = 34
    ws.row_dimensions[9].height = 22
    ws.row_dimensions[10].height = 22
    ws.freeze_panes = "A11"

    # ---- Row 11: 示例数据 ----
    examples = [
        (11, 1, "300866", "安克创新"),
        (11, 5, "AAPL", "Apple"),
        (11, 9, "00700", "腾讯控股"),
        (11, 13, "005930", "三星电子"),
    ]
    for row, code_col, code, name in examples:
        ws.cell(row=row, column=code_col, value=code).alignment = CENTER
        ws.cell(row=row, column=code_col + 1, value=name).alignment = CENTER

    for row in range(11, 32):
        ws.row_dimensions[row].height = 20
        for col in (1, 2, 5, 6, 9, 10, 13, 14):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = BORDER
            cell.alignment = CENTER if col in (1, 5, 9, 13) else LEFT


def build_wide_table(ws):
    """
    宽表 Sheet 的初始结构:
      A1=大类, B1=指标名称, R2 留空给 VBA 填日期, R1 C+ 给 VBA 填公司名
      列宽 A=30, B=40, C+=15.875 (但 C+ 只设默认列宽,VBA 写时会扩列)
      冻结 B3 (前 2 行表头 + 前 2 列锚定)
    """
    fill = PatternFill("solid", fgColor=DARK_BLUE)
    is_indicator = ws.title in ("A股_指标表", "美股_指标表", "港股_指标表", "韩股_指标表")

    ws["A1"] = "指标类型" if is_indicator else "大类"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = fill
    ws["A1"].alignment = CENTER
    ws["A1"].border = BORDER

    ws["B1"] = "指标名称"
    ws["B1"].font = HEADER_FONT
    ws["B1"].fill = fill
    ws["B1"].alignment = CENTER
    ws["B1"].border = BORDER

    if is_indicator:
        ws["C1"] = "英文指标名"
        ws["C1"].font = HEADER_FONT
        ws["C1"].fill = fill
        ws["C1"].alignment = CENTER
        ws["C1"].border = BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    if is_indicator:
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 34
    else:
        ws.column_dimensions["C"].width = 15.875
    ws.sheet_format.defaultColWidth = 15.875

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "D3" if is_indicator else "C3"


def build_cross_market_indicator_sheet(ws):
    """Phase 4g Step 2: 跨市场_指标表模板。"""
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 34
    ws.sheet_format.defaultColWidth = 15.875

    fill = PatternFill("solid", fgColor=DARK_BLUE)
    for cell_addr, txt in (("A1", "指标类型"), ("B1", "指标名称"), ("C1", "英文指标名")):
        cell = ws[cell_addr]
        cell.value = txt
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "D3"


def build_cross_market_statement_sheet(ws, statement_label):
    """Phase 4h Step 2: 跨市场 BS/IS/CF 合表模板。"""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.sheet_format.defaultColWidth = 15.875

    fill = PatternFill("solid", fgColor=DARK_BLUE)
    headers = (("A1", "大类"), ("B1", "指标名称"))
    for cell_addr, txt in headers:
        cell = ws[cell_addr]
        cell.value = txt
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER

    ws["A1"].comment = None
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "C3"


def build_corp_info(ws):
    headers = [
        ("A", "股票代码", 12),
        ("B", "股票简称", 14),
        ("C", "上市日期", 14),
        ("D", "所属行业", 24),
        ("E", "主营业务", 80),
    ]
    fill = PatternFill("solid", fgColor=DARK_BLUE)
    for col, name, width in headers:
        cell = ws[f"{col}1"]
        cell.value = name
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def build_diagnostic_sheet(ws, market_label="美股"):
    """
    Phase 4b-14a/4c: 抓取诊断 sheet 模板
    Row 1: 大标题 (合并 A1:K1, 深蓝白字)
    Row 2: 11 列表头 — 公司/报表/输出指标/状态/数据源/Taxonomy/命中字段/Unit/Score/匹配方式+备注/FX_Rate
      Row 3+: 由 VBA 写入(每次跑数后刷新)
      冻结 Row 2; 列宽 + 表头颜色与 install_modules.py._make_diagnostic_sheet
        和 模块_工具函数.bas.EnsureDiagnosticSheet() 三处保持一致
    """
    ws.sheet_state = "hidden"
    title_fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_font = Font(name="微软雅黑", size=12, bold=True, color=WHITE)

    # Row 1: 标题, 合并 A1:K1
    ws["A1"] = f"{market_label}抓取诊断 (每次跑数后自动刷新)"
    ws.merge_cells("A1:K1")
    for cell_addr in ["A1"]:
        c = ws[cell_addr]
        c.font = title_font
        c.fill = title_fill
        c.alignment = CENTER
        c.border = BORDER

    # Row 2: 11 列表头
    headers = ["公司", "报表", "输出指标", "状态", "数据源",
               "Taxonomy", "命中字段", "Unit", "Score", "匹配方式+备注", "FX_Rate"]
    header_font = Font(name="微软雅黑", size=10, bold=True, color=WHITE)
    for j, txt in enumerate(headers, start=1):
        col_letter = get_column_letter(j)
        c = ws[f"{col_letter}2"]
        c.value = txt
        c.font = header_font
        c.fill = title_fill
        c.alignment = CENTER
        c.border = BORDER

    widths = [14, 16, 30, 18, 18, 14, 42, 14, 10, 58, 12]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for row in range(1, 1001):
        ws.cell(row=row, column=1).number_format = "@"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20

    # 冻结 Row 2 (滚动时表头常驻)
    ws.freeze_panes = "A3"


def build_fx_sheet(ws):
    """
    Phase 4f Step 2: 汇率 sheet 模板
      Row 1: 8 列表头(深蓝白字)
      Row 2+: 由 VBA 模块_抓汇率 自动填充, 用户也可手填 override
      列宽: A=14 报告期 / B-G=14 数值 / H=40 备注
      冻结 A2; A 列文本格式 (防 yyyy-mm-dd 被 Excel 数字化)
    """
    ws.column_dimensions["A"].width = 14
    for letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[letter].width = 14
    ws.column_dimensions["H"].width = 40

    headers = ["报告期", "USDCNY期末", "USDCNY期均",
               "HKDCNY期末", "HKDCNY期均",
               "KRWCNY期末", "KRWCNY期均", "备注/override"]
    fill = PatternFill("solid", fgColor=DARK_BLUE)
    for j, txt in enumerate(headers, start=1):
        col = get_column_letter(j)
        cell = ws[f"{col}1"]
        cell.value = txt
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER

    # A 列文本格式 (防报告期被 Excel 数字化)
    for r in range(1, 200):
        ws.cell(row=r, column=1).number_format = "@"

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    # Drop the default sheet
    wb.remove(wb.active)

    # 1. 使用说明
    ws_intro = wb.create_sheet("使用说明")
    build_intro(ws_intro)

    # 2. 样本池
    ws_pool = wb.create_sheet("样本池")
    build_sample_pool(ws_pool)

    # ---- A 股 4 张表 (Phase 4b-4: 加 A股_ 前缀, 跟美股对称) ----
    for name in ["A股_资产负债表", "A股_利润表", "A股_现金流量表", "A股_指标表"]:
        ws = wb.create_sheet(name)
        build_wide_table(ws)

    # ---- Phase 4b: 美股 4 张表 (单位: 百万美元 / EPS 美元/股 / 加权股数 百万股) ----
    for name in ["美股_资产负债表", "美股_利润表", "美股_现金流量表", "美股_指标表"]:
        ws_us = wb.create_sheet(name)
        build_wide_table(ws_us)

    # ---- Phase 4b-14a: 美股_抓取诊断 sheet (放最后) ----
    ws_diag = wb.create_sheet("美股_抓取诊断")
    build_diagnostic_sheet(ws_diag, "美股")

    # ---- Phase 4c: 港股 4 张表 + 港股_抓取诊断 sheet ----
    for name in ["港股_资产负债表", "港股_利润表", "港股_现金流量表", "港股_指标表"]:
        ws_hk = wb.create_sheet(name)
        build_wide_table(ws_hk)

    ws_diag_hk = wb.create_sheet("港股_抓取诊断")
    build_diagnostic_sheet(ws_diag_hk, "港股")

    # ---- Phase 4d: 韩股 4 张表 + 韩股_抓取诊断 sheet ----
    for name in ["韩股_资产负债表", "韩股_利润表", "韩股_现金流量表", "韩股_指标表"]:
        ws_kr = wb.create_sheet(name)
        build_wide_table(ws_kr)

    ws_diag_kr = wb.create_sheet("韩股_抓取诊断")
    build_diagnostic_sheet(ws_diag_kr, "韩股")

    # ---- Phase 4h Step 2: 跨市场 BS/IS/CF 合并视图 ----
    for label in ("资产负债表", "利润表", "现金流量表"):
        ws_cross_stmt = wb.create_sheet(f"跨市场_{label}")
        build_cross_market_statement_sheet(ws_cross_stmt, label)

    # ---- Phase 4g Step 2: 跨市场指标合并视图 ----
    ws_cross = wb.create_sheet("跨市场_指标表")
    build_cross_market_indicator_sheet(ws_cross)

    # ---- Phase 4f Step 2: 汇率 sheet (跨市场共享缓存, 排序最末) ----
    ws_fx = wb.create_sheet("汇率")
    build_fx_sheet(ws_fx)

    # 默认打开时显示样本池
    wb.active = wb.sheetnames.index("样本池")

    wb.save(OUT_PATH)
    print(f"Generated: {OUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
