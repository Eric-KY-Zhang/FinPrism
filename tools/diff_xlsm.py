"""
Compare statement data cells between two .xlsm workbooks.

Usage:
    py tools/diff_xlsm.py [new.xlsm] [baseline.xlsm]

Default:
    new      = ../上市公司财务数据查询.xlsm
    baseline = ../archive/上市公司财务数据查询_4b14a_baseline_20260503.xlsm

Scope:
    Six statement sheets, rows 3+ only:
      A股_资产负债表 / A股_利润表 / A股_现金流量表
      美股_资产负债表 / 美股_利润表 / 美股_现金流量表

Indicator sheets, diagnostic sheets, sample pool, and instructions are skipped.
The script exits with status 1 when any mismatch is found.
"""

from __future__ import annotations

import argparse
import math
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_NEW = ROOT / "上市公司财务数据查询.xlsm"
DEFAULT_BASELINE = ROOT / "archive" / "上市公司财务数据查询_4b14a_baseline_20260503.xlsm"

SHEETS = [
    "A股_资产负债表",
    "A股_利润表",
    "A股_现金流量表",
    "美股_资产负债表",
    "美股_利润表",
    "美股_现金流量表",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("new", nargs="?", type=Path, default=DEFAULT_NEW)
    parser.add_argument("baseline", nargs="?", type=Path, default=DEFAULT_BASELINE)
    parser.add_argument("--max-mismatches", type=int, default=50)
    parser.add_argument("--tolerance", type=float, default=1e-9)
    return parser.parse_args()


def comparable(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        return None if s == "" else s
    return value


def values_equal(left: Any, right: Any, tolerance: float) -> bool:
    left = comparable(left)
    right = comparable(right)
    if left is None and right is None:
        return True
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        if math.isnan(float(left)) and math.isnan(float(right)):
            return True
        return abs(float(left) - float(right)) <= tolerance
    return left == right


def sheet_values(ws, min_row: int = 3) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, values_only=True), start=min_row):
        for col_idx, raw_value in enumerate(row, start=1):
            value = comparable(raw_value)
            if value is not None:
                values[(row_idx, col_idx)] = value
    return values


def compare_sheet(ws_new, ws_base, max_mismatches: int, tolerance: float) -> list[str]:
    new_values = sheet_values(ws_new)
    base_values = sheet_values(ws_base)
    mismatches: list[str] = []

    for row, col in sorted(set(new_values) | set(base_values)):
        new_value = new_values.get((row, col))
        base_value = base_values.get((row, col))
        if not values_equal(new_value, base_value, tolerance):
            addr = f"{get_column_letter(col)}{row}"
            mismatches.append(
                f"{addr}: new={comparable(new_value)!r} baseline={comparable(base_value)!r}"
            )
            if len(mismatches) >= max_mismatches:
                return mismatches
    return mismatches


def main() -> int:
    args = parse_args()
    if not args.new.exists():
        print(f"Missing new workbook: {args.new}", file=sys.stderr)
        return 2
    if not args.baseline.exists():
        print(f"Missing baseline workbook: {args.baseline}", file=sys.stderr)
        return 2

    wb_new = load_workbook(args.new, data_only=True, read_only=True, keep_vba=True)
    wb_base = load_workbook(args.baseline, data_only=True, read_only=True, keep_vba=True)

    total = 0
    try:
        for sheet_name in SHEETS:
            if sheet_name not in wb_new.sheetnames:
                print(f"{sheet_name}: missing in new workbook")
                total += 1
                continue
            if sheet_name not in wb_base.sheetnames:
                print(f"{sheet_name}: missing in baseline workbook")
                total += 1
                continue
            mismatches = compare_sheet(
                wb_new[sheet_name],
                wb_base[sheet_name],
                args.max_mismatches,
                args.tolerance,
            )
            total += len(mismatches)
            if mismatches:
                print(f"{sheet_name}: {len(mismatches)} mismatches (showing up to {args.max_mismatches})")
                for line in mismatches:
                    print(f"  {line}")
            else:
                print(f"{sheet_name}: 0 mismatches")
    finally:
        wb_new.close()
        wb_base.close()

    if total:
        print(f"TOTAL: {total} mismatches")
        return 1
    print("TOTAL: 0 mismatches")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
